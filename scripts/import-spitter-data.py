"""
Import spitter/emitter data from OrchardInfo.xlsx into orchards table.

Populates: spitter_type_id, spitters_per_ha, spitter_tree_ratio, water_meter
Maps SpitterType integer (from Excel) to spitter_types table via name lookup.

Usage:
  python scripts/import-spitter-data.py
"""

import os
import sys
import logging

sys.path.insert(0, 'C:/Python314/Lib/site-packages')

import openpyxl
import requests

logging.basicConfig(level=logging.INFO, format='%(asctime)s %(levelname)s %(message)s', datefmt='%Y-%m-%d %H:%M:%S')
log = logging.getLogger(__name__)

SUPABASE_URL = 'https://agktzdeskpyevurhabpg.supabase.co'
ORG_ID = '93d1760e-a484-4379-95fb-6cad294e2191'

ORCHARD_INFO_PATH = r'C:\Users\rickus.MOUTONSVALLEY\OneDrive - Moutons Valley Trust\Attachments\OrchardInfo.xlsx'
SPITTER_TYPES_PATH = r'C:\Users\rickus.MOUTONSVALLEY\OneDrive - Moutons Valley Trust\Attachments\SpitterTypes.xlsx'

def _read_env_local():
    env_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), '..', '.env.local')
    try:
        with open(env_path) as f:
            for line in f:
                line = line.strip()
                if line.startswith('SUPABASE_SERVICE_ROLE_KEY='):
                    return line.split('=', 1)[1].strip()
    except Exception:
        pass
    return ''

SERVICE_KEY = _read_env_local()
HEADERS = {
    'apikey': SERVICE_KEY,
    'Authorization': f'Bearer {SERVICE_KEY}',
    'Content-Type': 'application/json',
    'Prefer': 'return=minimal',
}


def main():
    # 1. Load spitter types from Excel (ID → name mapping)
    log.info("Loading SpitterTypes from %s", SPITTER_TYPES_PATH)
    wb_st = openpyxl.load_workbook(SPITTER_TYPES_PATH, read_only=True, data_only=True)
    ws_st = wb_st[wb_st.sheetnames[0]]
    excel_spitter_map = {}  # Excel ID (int) → name
    for row in ws_st.iter_rows(min_row=2, values_only=True):
        if row[0] is not None and row[1]:
            excel_spitter_map[int(row[0])] = row[1]
    wb_st.close()
    log.info("Excel spitter types: %s", excel_spitter_map)

    # 2. Fetch spitter_types from Supabase (name → uuid)
    resp = requests.get(
        f'{SUPABASE_URL}/rest/v1/spitter_types?select=id,name',
        headers={'apikey': SERVICE_KEY, 'Authorization': f'Bearer {SERVICE_KEY}'},
        timeout=30,
    )
    resp.raise_for_status()
    db_spitter_map = {row['name']: row['id'] for row in resp.json()}
    log.info("DB spitter types: %d", len(db_spitter_map))

    # Build Excel ID → DB UUID
    spitter_id_map = {}
    for excel_id, name in excel_spitter_map.items():
        if name in db_spitter_map:
            spitter_id_map[excel_id] = db_spitter_map[name]
        else:
            log.warning("Spitter type '%s' (Excel ID %d) not in DB", name, excel_id)
    log.info("Mapped %d spitter types", len(spitter_id_map))

    # 3. Fetch orchards from Supabase (legacy_id → id)
    resp = requests.get(
        f'{SUPABASE_URL}/rest/v1/orchards?organisation_id=eq.{ORG_ID}&legacy_id=not.is.null&select=id,legacy_id',
        headers={'apikey': SERVICE_KEY, 'Authorization': f'Bearer {SERVICE_KEY}'},
        timeout=30,
    )
    resp.raise_for_status()
    orchard_map = {row['legacy_id']: row['id'] for row in resp.json()}
    log.info("Orchards with legacy_id: %d", len(orchard_map))

    # 4. Parse OrchardInfo.xlsx
    log.info("Loading OrchardInfo from %s", ORCHARD_INFO_PATH)
    wb = openpyxl.load_workbook(ORCHARD_INFO_PATH, read_only=True, data_only=True)
    ws = wb['OrchardInfo']
    headers_row = list(next(ws.iter_rows(min_row=1, max_row=1, values_only=True)))

    updated = 0
    skipped = 0

    for row in ws.iter_rows(min_row=2, values_only=True):
        vals = list(row)
        d = dict(zip(headers_row, vals))

        legacy_id = d.get('OrchardID')
        if legacy_id is None:
            skipped += 1
            continue
        legacy_id = int(legacy_id)

        orchard_uuid = orchard_map.get(legacy_id)
        if not orchard_uuid:
            skipped += 1
            continue

        spitter_excel_id = d.get('SpitterType')
        spitter_type_id = spitter_id_map.get(int(spitter_excel_id)) if spitter_excel_id else None

        spitters_per_ha = d.get('Spitters/ha')
        spitter_tree_ratio = d.get('Spitter / Tree Ratio')
        water_meter_val = d.get('WaterMeter')

        # Build update payload (only non-null fields)
        patch = {}
        if spitter_type_id:
            patch['spitter_type_id'] = spitter_type_id
        if spitters_per_ha is not None:
            patch['spitters_per_ha'] = round(float(spitters_per_ha), 2)
        if spitter_tree_ratio is not None:
            patch['spitter_tree_ratio'] = round(float(spitter_tree_ratio), 2)
        if water_meter_val is not None:
            patch['water_meter'] = str(water_meter_val)

        if not patch:
            skipped += 1
            continue

        resp = requests.patch(
            f'{SUPABASE_URL}/rest/v1/orchards?id=eq.{orchard_uuid}',
            headers=HEADERS,
            json=patch,
            timeout=30,
        )
        if resp.status_code >= 400:
            log.warning("Failed to update orchard %s (legacy %d): %d %s",
                        orchard_uuid, legacy_id, resp.status_code, resp.text[:200])
        else:
            updated += 1

    wb.close()
    log.info("Updated %d orchards, skipped %d", updated, skipped)


if __name__ == '__main__':
    main()
